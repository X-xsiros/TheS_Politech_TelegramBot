import logging
import os
import datetime
import threading
import random, requests
import openai
import openpyxl
from openpyxl import Workbook

from aiogram import Bot, Dispatcher, types, executor
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from dotenv import load_dotenv
from sqlalchemy.orm import Session
from apscheduler.schedulers.asyncio import AsyncIOScheduler

import data.keyboards as kb  # do your keyboards HERE!
from data.db_session import global_init, create_session
from data.groups import Groups
from data.homework import Homework

if os.path.exists('.env'):
    load_dotenv('.env')  # call AlkoKovad if you don't have one

BOT_KEY = os.getenv('API_TOKEN')
STORAGE_ID = -1001461174439
stpass = 'feelgod'

lock = threading.Lock()
storage = MemoryStorage()
logging.basicConfig(level=logging.INFO)
bot = Bot(token=BOT_KEY)
dp = Dispatcher(bot, storage=storage)
scheduler = AsyncIOScheduler()


class HomeworkSendState(StatesGroup):
    homework_files_state = State()
    homework_files_state2 = State()
    homework_files_state_download = State()


class DeadlineWriteState(StatesGroup):
    deadline_date = State()
    deadline_text = State()
class Rootstate(StatesGroup):
    root1 = State()
    root2 = State()
class Change(StatesGroup):
    change1 = State()
class Gnumber(StatesGroup):
    gnumber1 = State()
class Attend(StatesGroup):
    attend1 = State()
    attend2 = State()
    attend3 = State()

scheduler.start()


@dp.message_handler(commands=['start', 'menu'])
async def send_menu(message: types.Message):
    """Main menu entry command and nothing else."""

    db_sess = create_session()
    group_ids = [chat.group_chat_id for chat in db_sess.query(Groups).all()]

    if str(message.chat.id) in group_ids:
        await message.answer('В общих чатах эти команды недоступны, '
                             'потому что для меня они всё равно что интимны.')
    elif int(message.chat.id) > 0:
        await message.answer('Меню', reply_markup=kb.menu(message.message_id))
    else:
        group = Groups()
        group.group_chat_id = message.chat.id
        db_sess.add(group)
        await bot.send_message(message.chat.id,'Обновления: Исправлены критические ошибки \n ВНИМАНИЕ всем нужно заново авторизоватся')
        db_sess.commit()

@dp.message_handler(commands=['cat'])
async def randomcat(message: types.Message):
    r = requests.get('https://cataas.com/cat')
    await bot.send_photo(message.chat.id, r.content, disable_notification=True)


@dp.message_handler(commands=['capy'])
async def randomcapy(message: types.Message):
    r = requests.get('https://api.capy.lol/v1/capybara/')
    await bot.send_photo(message.chat.id, r.content, disable_notification=True)

@dp.message_handler(commands=['ai'])
async def genaratepicture(message: types.Message):
    response = openai.Image.create(
        api_key=os.getenv('AI_TOKEN'),
        prompt=message.text,
        n=1,
        size="1024x1024",
    )
    imageurl = response['data'][0]['url']
    await bot.send_photo(message.chat.id, imageurl)

@dp.callback_query_handler()
async def subject_menu(callback_query: types.CallbackQuery,state: FSMContext):
    """Subject menu is a hub for a homework, docs and literature sections;
    There is a data format for reply markups which looks like:
    Section-subject-kind_of_docs"""

    db_sess: Session = create_session()


    info_type, bot_msg = callback_query.data.split()[0], int(callback_query.data.split()[1]) + 1
    # this string could fuck up our bot one day, search "last_message" or smth like that command plz...
    await bot.delete_message(chat_id=callback_query.from_user.id, message_id=bot_msg)
    await bot.answer_callback_query(callback_query.id)  # have no fucking clue what is this for

    if info_type == 'menu':
        await bot.send_message(callback_query.from_user.id, 'Меню', reply_markup=kb.menu(bot_msg))

        return  # menu exit
    if 'get' in info_type:

        db_sess = create_session()
        deaddata = info_type[-10:]

        user = '@' + callback_query.from_user.username
        group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()
        homework_download = db_sess.query(Homework).filter(Homework.group_chat_id == group.group_chat_id,
                                                           Homework.deadline == deaddata,Homework.subject == info_type.split('-')[1])

        for i in homework_download:
            await bot.forward_message(callback_query.from_user.id, STORAGE_ID, i.homework_id)

        db_sess.commit()
        return
    if 'ns' in info_type:
        db_sess = create_session()
        user = info_type.split('-')[-1]
        group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()
        members = group.members
        index = members.split(' ').index(user)
        await bot.send_message(callback_query.from_user.id,f"Текущий счет {group.score.split(' ')[index]}, сколко добавить?(Напиши +\-колво баллов)")
        async with state.proxy() as data:
            data['euser'] = user
        await Rootstate.root2.set()
        db_sess.commit()
        return
    if 'my' in info_type:
        db_sess = create_session()
        user = callback_query.from_user.username
        group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()
        members = group.members
        index = members.split(' ').index('@' + user)
        score = group.score.split(' ')[index]
        await bot.send_message(callback_query.from_user.id, f'Баллов у {user},{score}')
        db_sess.commit()
        return
    if 'edit' in info_type:
        db_sess = create_session()
        user = callback_query.from_user.username
        group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()

        if user in group.root:
            members = group.members

            names = members.split(' ')
            await bot.send_message(callback_query.from_user.id, 'Выбери ученика',
                                   reply_markup=kb.score_edit(bot_msg, info_type, names))

        else:
            await bot.send_message(callback_query.from_user.id, 'Ты не староста')
        db_sess.commit()
        return

    if 'all' in info_type:
        db_sess = create_session()
        user = callback_query.from_user.username
        group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()

        if user in group.root:
            members = group.members.split(' ')
            score = group.score.split(' ')
            for i in range(len(members) - 1):
                await bot.send_message(callback_query.from_user.id,f'{members[i]} {score[i]}')
        else:
            await bot.send_message(callback_query.from_user.id, 'Ты не староста')
        return

    if 'check' in info_type:
        if 'starbtn' in info_type:
            db_sess = create_session()
            user = callback_query.from_user.username
            group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()
            if user in group.root:
                await bot.send_message(callback_query.from_user.id, 'Нажми чтобы установить локацию', reply_markup=kb.locate())
                async with state.proxy() as data:
                    data['subject'] = info_type.split('-')[1]
                    data['group'] = group.group_number
                await Attend.attend1.set()
            else:
                await bot.send_message(callback_query.from_user.id, 'Ты не староста')
            return
        elif 'studbtn' in info_type:
            db_sess = create_session()
            user = callback_query.from_user.username
            group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()

            await bot.send_message(callback_query.from_user.id, 'Нажми чтобы отправить локацию',
                                       reply_markup=kb.locate())
            async with state.proxy() as data:
                data['subject'] = info_type.split('-')[1]
                data['group'] = group.group_number
                data['user'] = user
            await Attend.attend2.set()

        else:
            await bot.send_message(callback_query.from_user.id, 'Выбери действие',
                                   reply_markup=kb.check(bot_msg, info_type))
        return

    if 'docs-' in info_type:
        if len(info_type.split('-')) == 3:
            await bot.send_message(callback_query.from_user.id, 'Выбери действие',
                                   reply_markup=kb.up_down__load_files(bot_msg, info_type))
            return

        if len(info_type.split('-')) == 4:

            if 'upload' in info_type:
                await bot.send_message(callback_query.from_user.id,
                                       'Укажи дату в формате YYYY-MM-DD и приложи файлы')
                async with state.proxy() as data:
                    data['subject'] = info_type.split('-')[1]
                await HomeworkSendState.homework_files_state.set()
                return  # upload homework entry
            if 'download' in info_type:
                user = callback_query.from_user.username
                group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()
                homework_list = db_sess.query(Homework).filter(Homework.subject == info_type.split('-')[1],Homework.group_chat_id == group.group_chat_id).all()
                dates = [homework.deadline for homework in homework_list]
                await bot.send_message(callback_query.from_user.id, 'Ваше ДЗ по дедлайнам хозяин',
                                       reply_markup=kb.homework_for_dates(bot_msg, info_type, dates))

                return  # download homework entry

        await bot.send_message(callback_query.from_user.id, 'Выбери материалы',
                               reply_markup=kb.materials_by_subject(bot_msg, info_type))

        return  # materials section entry
    if 'score' in info_type:
        await bot.send_message(callback_query.from_user.id, 'Выбери действие',reply_markup=kb.score_chose(bot_msg, info_type))

        return

    user = callback_query.from_user.username
    subjects_list = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()
    subjects = subjects_list.subjects.split('-')
    await bot.send_message(callback_query.from_user.id, 'Выбери предмет',
                           reply_markup=kb.subjects(bot_msg, info_type.split('-')[0],subjects))  # subject entry

@dp.message_handler(state=HomeworkSendState.homework_files_state)
async def upload_homework(message: types.message, state: FSMContext):
    if len(message.text) == 10 and type(int(message.text.split('-')[0])) == int:
        await state.update_data(deadline_date=message.text)
        await bot.send_message(message.chat.id, 'Отправь файлы')
        await HomeworkSendState.homework_files_state2.set()
    else:
        await message.reply(f"Неверный формат{len(message.text.split('-'))}{message.text.split('-')}")


@dp.message_handler(state=HomeworkSendState.homework_files_state2, content_types=['text', 'photo', 'document'])
async def upload_homework2(message: types.message, state: FSMContext):
    async with state.proxy() as data:
        sub = data['subject']
    db_sess = create_session()
    user = '@' + message.from_user.username
    group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()

    await state.update_data(homework=message)
    data = await state.get_data()
    deadline_date = datetime.datetime.strptime(data['deadline_date'], '%Y-%m-%d')

    try:

        if data['homework'].document is not None:
            msg = await bot.send_document(STORAGE_ID, document=data['homework'].document.file_id, caption= deadline_date)
            db_sess.add(
                Homework(group_chat_id=group.group_chat_id, deadline=deadline_date, homework_id=msg["message_id"],
                         some_text='Смотри дз',subject= sub))
        elif data['homework'].text is not None:
            msg = await bot.send_message(STORAGE_ID, f"{deadline_date},{data['homework'].text},")
            db_sess.add(
                Homework(group_chat_id=group.group_chat_id, deadline=deadline_date, homework_id=msg["message_id"],
                         some_text='Смотри дз',subject= sub))

        elif data['homework'].photo is not None:
            msg = await bot.send_photo(STORAGE_ID, photo=data['homework'].photo[-1].file_id, caption= deadline_date)
            db_sess.add(
                Homework(group_chat_id=group.group_chat_id, deadline=deadline_date, homework_id=msg["message_id"],
                         some_text='Смотри дз',subject= sub))

        await state.finish()
        db_sess.commit()
        await bot.send_message(message.chat.id, 'успешно добавлено')
    except Exception as e:
        await bot.send_message(message.chat.id, f'Ошибка, сообщите админу {e}')
    return


@dp.message_handler(commands=['alldeadlines'])
async def alldeadlines(message: types.Message):
    db_sess = create_session()
    user = '@' + message.from_user.username
    group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()
    deadlines = db_sess.query(Homework).filter(
        Homework.group_chat_id == group.group_chat_id).all()

    if deadlines == []:
        await bot.send_message(message.chat.id, f'Дедлайнов нет для группы:{group.group_chat_id}')
    else:
        deadlines = db_sess.query(Homework).filter(
            Homework.group_chat_id == group.group_chat_id)
        await bot.send_message(message.chat.id, 'Дедлыйны:')
        for c in deadlines:
            await bot.send_message(message.chat.id,
                                   f'\n Дата {c.deadline}\n Id Дз {c.homework_id}\n Предмет {c.subject} \n Суть  {c.some_text}')

    db_sess.commit()


@dp.message_handler(commands=['adddeadline'])
async def adddeadline(message: types.Message):
    await bot.send_message(message.chat.id, 'Введи дедлайн в формате YYYY-MM-DD')
    await DeadlineWriteState.deadline_date.set()

@dp.message_handler(state=Attend.attend1, content_types=['location'],is_reply=True)
async def lessonc(message: types.Message,state: FSMContext):
    lat = message.location.latitude
    lon = message.location.longitude
    data = await state.get_data()
    reply = "latitude:  {}\nlongitude: {}".format(lat, lon)
    nless(data['group'],data['subject'],lat,lon)
    await message.answer(reply, reply_markup=types.ReplyKeyboardRemove())
    await state.finish()

@dp.message_handler(state= Attend.attend2,content_types=['location'],is_reply=True)
async def attend(message: types.Message,state:FSMContext):
    lat = message.location.latitude
    lon = message.location.longitude
    reply = "latitude:  {}\nlongitude: {}".format(lat, lon)
    data = await state.get_data()
    group = data['group']
    sub = data['subject']
    if (not os.path.isfile(f"attendance/{group}.xlsx")):
        await  message.reply('Ты слишком далеко или пары нет')
        await state.finish()
    wb = openpyxl.load_workbook(f"attendance/{group}.xlsx")
    ws = wb[f'{sub}']

    await message.answer(reply, reply_markup=types.ReplyKeyboardRemove())
    await state.finish()
def attless(group,sub,lat,lon):
    if (os.path.isfile(f"attendance/{group}.xlsx")):
        wb = openpyxl.load_workbook(f"attendance/{group}.xlsx")
    else:
        return 1
    ws = wb[f'{sub}']
    if(abs((ws.cell(row=1, column=1).value + ws.cell(row=1, column=2).value) - (lat+lon)) >25 ):
        wb.save(f"attendance/{group}.xlsx")
        return 1
    elif(ws.cell(row=1, column=3).value ==  1):
        return 2
    else:
        return 3

def nless(group,sub,lat,lon):

    if (os.path.isfile(f"attendance/{group}.xlsx")):
        wb = openpyxl.load_workbook(f"attendance/{group}.xlsx")
    else:
        wb = Workbook()
    if f'{sub}' in wb.sheetnames:
        ws = wb[f'{sub}']
    else:
        wb.create_sheet(f'{sub}')
        ws = wb[f'{sub}']
        ws.cell(row=1, column=3).value = 0
        ws.cell(row=1, column=4).value = 0
    ws.cell(row=1, column=1).value = lat
    ws.cell(row=1, column=2).value = lon
    ws.cell(row=1, column=3).value +=  1
    ws.cell(row=2, column= 2 + ws.cell(row=1, column=3).value).value = datetime.date.today()
    wb.save(f"attendance/{group}.xlsx")

@dp.message_handler(state=DeadlineWriteState.deadline_date)
async def get_ddate(message: types.Message, state: FSMContext):
    if len(message.text) == 10 and type(int(message.text.split('-')[0])) == int:
        await state.update_data(deadline_date=message.text)
        await bot.send_message(message.chat.id, 'Введи суть деделайна')
        await DeadlineWriteState.deadline_text.set()
    else:
        await message.reply(f"Неверный формат{len(message.text.split('-'))}{message.text.split('-')}")


@dp.message_handler(state=DeadlineWriteState.deadline_text)
async def get_dtext(message: types.Message, state: FSMContext):
    await state.update_data(deadline_text=message.text)
    data = await state.get_data()
    db_sess = create_session()
    user = '@' + message.from_user.username
    group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()
    deadline_date = datetime.datetime.strptime(data['deadline_date'], '%Y-%m-%d')
    msg = await bot.send_message(STORAGE_ID, f"{data['deadline_text']}")
    try:
        db_sess.add(Homework(group_chat_id=group.group_chat_id, deadline=deadline_date, homework_id=msg['message_id'],
                             some_text=data['deadline_text'],subject= 'Deadline'))
        print(type(group.group_chat_id), type(deadline_date), type(os.urandom(5)))
    except Exception as e:
        await message.reply(f'{e}')
    await message.reply('Дедлайн успешно создан')
    db_sess.commit()
    await state.finish()


@dp.message_handler(commands=['deadline'])
async def deadline(message: types.Message):
    db_sess = create_session()
    tomorrow = datetime.date.today() + datetime.timedelta(days=1)
    user = '@' + message.from_user.username

    group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()

    deadlines = db_sess.query(Homework).filter(
        Homework.group_chat_id == group.group_chat_id, Homework.deadline == tomorrow).all()

    if deadlines == []:
        await bot.send_message(message.chat.id, f'Дедлайнов на завтра нет для группы:{group.group_chat_id}')
    else:
        deadlines = db_sess.query(Homework).filter(
            Homework.group_chat_id == group.group_chat_id, Homework.deadline == tomorrow)
        for c in deadlines:
            await bot.send_message(message.chat.id,
                                   f'Дедлайны на завтра\n Дата {c.deadline}\n Id Дз {c.homework_id}\n Предмет {c.subject} \n Суть  {c.some_text}')

    db_sess.commit()


@dp.message_handler(commands=['help'])
async def help(message: types.Message):
    await bot.send_message(message.chat.id,'Команды:\n /help\n /deadline\n /alldeadlines\n /adddeadline\n /alarm\n /start\n /root')

@dp.message_handler(commands=['root'])
async def root(message: types.Message):
    await bot.send_message(message.chat.id,'Введите пароль старосты')
    await Rootstate.root1.set()

@dp.message_handler(state= Rootstate.root1)
async def root1(message: types.Message,state: FSMContext):
    db_sess = create_session()
    password = message.text
    user = message.from_user.username
    group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()
    if password == stpass:
        if user in group.root:
            await bot.send_message(message.from_user.id,'Ты уже староста')
        else:


            group.root += user + " "
            await bot.send_message(message.chat.id, 'Преветсвую, староста, чем могу служить?')
    else:
        await bot.send_message(message.chat.id,'ПНХ')
    await state.finish()
    db_sess.commit()

@dp.message_handler(state= Rootstate.root2)
async def edit_score(message: types.Message,state:FSMContext):
    msg = message.text
    if msg[0]=='-' or msg[0]=='+':
        db_sess = create_session()
        async with state.proxy() as data:
            user = data['euser']
        group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()
        members = group.members
        members = members.split(' ')
        index = members.index(user)
        new = int(group.score.split(' ')[index]) + int(msg)
        score = group.score.split(' ')
        score[index] = new
        group.score = ' '.join(map(str,score))
        await bot.send_message(message.chat.id,'Успешно изменено')
        db_sess.commit()
    else:
        await bot.send_message(message.chat.id,"Неверный формат")
    await state.finish()
    return

@dp.message_handler(commands=['alarm'])
async def alarm_command(message: types.Message):
    """Just a command which doing that king of thing in chat: @all"""

    db_sess = create_session()
    group = db_sess.query(Groups).filter(Groups.group_chat_id == message.chat.id).first()
    await bot.send_message(message.chat.id, group.members)

@dp.message_handler(commands=['change'])
async def change(message: types.Message):
    db_sess = create_session()
    user = message.from_user.username
    group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()
    if user in group.root:
        await bot.send_message(message.chat.id,"Отправь новые предметы")
        db_sess.commit()
        await Change.change1.set()
    else:
        await bot.send_message(message.chat.id, 'Ты не староста')

    db_sess.commit()
@dp.message_handler(state= Change.change1)
async def change2(message: types.Message,state: FSMContext):
    db_sess = create_session()
    subjects = message.text
    user = message.from_user.username
    group = db_sess.query(Groups).filter(Groups.members.like(f'%{user}%')).first()
    group.subjects = subjects
    db_sess.commit()
    await state.finish()

@dp.message_handler(commands=['gnumber'])
async def gnumber(message: types.Message):
    if (message.chat.id < 0):
        db_sess = create_session()
        user = message.from_user.username
        group = db_sess.query(Groups).filter(Groups.group_chat_id == message.chat.id).first()
        if user in group.root:
            await bot.send_message(message.chat.id,"Введи номер группы")
            db_sess.commit()
            await Gnumber.gnumber1.set()
        else:
            await bot.send_message(message.chat.id, 'Ты не староста')
        db_sess.commit()
    else:
        await bot.send_message(message.chat.id,"Доступно только в чате группы")

@dp.message_handler(state=Gnumber.gnumber1)
async def gnumber1(message: types.Message,state: FSMContext):
    db_sess = create_session()
    gnumber = message.text
    group = db_sess.query(Groups).filter(Groups.group_chat_id == message.chat.id).first()
    group.group_number = gnumber
    db_sess.commit()
    await bot.send_message(message.chat.id, "Спасибо")
    await state.finish()

@dp.message_handler()
async def usernames_capture(message: types.Message):
    """This function captures usernames from messages in group chats,
    then it makes or opens a new group usernames file, next it checks
    if the username is already in, if not the function write it down."""

    db_sess = create_session()
    if db_sess.query(Groups).filter(Groups.group_chat_id == message.chat.id).first():
        group = db_sess.query(Groups).filter(Groups.group_chat_id == message.chat.id).first()
        if message.from_user.username == None:
            await bot.send_message(message.chat.id, 'Пидор, быстро сука завел юзернэйм! АТОБАН')

        elif '@' + message.from_user.username + ' ' not in group.members and message.from_user.username != None:
            group.members += '@' + message.from_user.username + ' '
            group.score += ' 0'

    db_sess.commit()

# @dp.message_handler(commands=['remove'])
# async def remove(message: types.Message):
#     db_sess = create_session()
#     group = db_sess.query(Groups).filter(Groups.group_chat_id == message.chat.id).first()
#     message.text.split(' ')

def search_user_in_groups(username):
    db_sess = create_session()
    groups = db_sess.query(Groups).all()
    for group in groups:
        if username in group.members:
            return group.group_chat_id


async def reminder(dp):
    db_sess = create_session()
    tomorrow = datetime.date.today() + datetime.timedelta(days=1)
    groups = db_sess.query(Groups)
    for group in groups:
        deadlines = db_sess.query(Homework).filter(
            Homework.group_chat_id == group.group_chat_id, Homework.deadline == tomorrow).all()
        if deadlines == []:
            await dp.bot.send_message(group.group_chat_id,"Дедлайнов нет")
        else:
            deadlines = db_sess.query(Homework).filter(
                Homework.group_chat_id == group.group_chat_id, Homework.deadline == tomorrow)
            for c in deadlines:
                await dp.bot.send_message(group.group_chat_id,
                                          f'Дедлайны на завтра\n Дата {c.deadline}\n Id Дз {c.homework_id}\n Суть  {c.some_text}')
    houmwork = db_sess.query(Homework).filter(Homework.deadline < datetime.date.today()).all()
    try:
        if houmwork != []:
            houmwork = db_sess.query(Homework).filter(Homework.deadline < datetime.date.today())
            for houm in houmwork:
                db_sess.delete(houm)
    except Exception as e:
        await dp.bot.send_message(STORAGE_ID,f"{e}")

    db_sess.commit()


scheduler.add_job(reminder, "cron", hour= 14, minute= 1, args=(dp,))


def db_connect():
    """DataBase connection"""

    db_name = 'db/db.sqlite'
    global_init(db_name)


if __name__ == '__main__':
    db_connect()
    executor.start_polling(dp, skip_updates=True)
